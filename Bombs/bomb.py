from openpyxl import load_workbook
import random as rd
workbook = load_workbook(filename="bomb.xlsx") #ดึงข้อมูลจากไฟล์
workbook.sheetnames
sheet = workbook.active
sheet.title
char = ["A","B","C","D","E","F","G","H","I","J","K","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
#player = int(input("how much player you want to play: "))

for i in range(len(char)):
    for y in range(1,26):
        sheet[char[i]+str(y)] = ""
for t in range(len(char)):
    for u in range(2):
        target = rd.randint(1,25)
        sheet[char[t]+str(target)] = "BOMB"
#สุ่มระเบิด 2 ลูกในแต่ละแถว
workbook.save(filename="bomb.xlsx")
num = 0
while True:
    choice = input("Player choose your target: ")
    print(choice)
    if sheet[str(choice)].value == "BOMB":
        print("gameover")
        break
    elif sheet[str(choice)].value != "BOMB":
        num +=1
        continue

